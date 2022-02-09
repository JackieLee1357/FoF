#!/E:\PythonProjects\pythonProject2 python3.8
# -*- coding: utf-8 -*-
# ---
# @Software: PyCharm
# @File: notebook.py
# @Author: Jackie Lee
# @Institution: Wuxi, Jiangsu, China
# @E-mail: Yuan_li5928@jabil.com
# @Site: 
# @Time: 8月 12, 2021
# ---


# pip3 install numpy --target=E:\Anaconda3\envs\pyinstaller  # 安装包至指定文件夹
# 创建虚拟环境
# conda create - n aotu python = 3.6
#
# # 激活虚拟环境
# conda activate aotu
#
# # Pyinstaller打包
# Pyinstaller - F - w - i apple.ico py_word.py
#
#
# conda remove -n aotu--all
# Ctrl + B    # 查看源码
# path = os.path.abspath(os.path.dirname(sys.argv [0]))    # 获取当前执行文件夹路径
# config.read(path0, encoding="utf-8-sig")  # 导出配置文件,解决文档格式为utf-8-bom问题
# https://blog.csdn.net/weixin_38753213/article/details/111399147 读取和写入excel图片
# people_email = people_email.split(",")   # 字符串根据","进行分割，转换为list
# dates = [''.join(i) for i in dates]  # 元组转化为列表
# plt.__dict__   # 查看方法
# pandas.set_option('display.max_columns', None)  # 显示所有列
# df.reset_index(level=0, inplace=True)    # df 索引转列
# df.reset_index(inplace=True, drop=False)# drop 是否把index列丢弃
# df['index1'] = df.index    # df 索引转列


# df3 = pd.DataFrame({'key1': ['K0', 'K0', 'K1', 'K2'],
#                     'key2': ['K0', 'K1', 'K0', 'K1'],
#                     'A': ['A0', 'A1', 'A2', 'A3'],
#                     'B': ['B0', 'B1', 'B2', 'B3']})
# df4 = pd.DataFrame({'key1': ['K0', 'K1', 'K1', 'K2'],
#                     'key2': ['K0', 'K0', 'K0', 'K0'],
#                     'C': ['C0', 'C1', 'C2', 'C3'],
#                     'D': ['D0', 'D1', 'D2', 'D3']})
# # 多个链接键,必须key1和key2都相同才合并
# print(pd.merge(df3, df4, on=['key1','key2']))           # pandas链接 join in

# s1 = pd.Series([1,2,3])
# s2 = pd.Series([2,3,4])
# print(pd.concat([s1,s2]))       # pandas连接相加
# print('-----')
# # 默认axis=0，行+行

#  sq = s.unique()          # 去重
# s = pd.Series([1,1,1,1,2,2,2,3,4,5,5,5,5])
# print(s.duplicated())# 判断是否重复
# print(s[s.duplicated() == False])# 通过布尔判断，得到不重复的值
# # drop.duplicates移除重复
# # inplace参数：是否替换原值，默认False
# s_re = s.drop_duplicates()
# Ctrl+Alt+L   格式化json
# '{:,}'.format(b)  数字加千位分隔符
#  '{:.1%}'.format(b)   百分数格式
# https://www.cnblogs.com/qinchao0317/p/10699717.html   字符串格式化
# facaDate = sheetData[sheetData['漏失率'] > 0.001]   # 按条件筛选数据，返回数据
# facaDate = sheetData[sheetData['漏失率'] > 0.001,]   # 按条件筛选数据，返回布尔值
# facaDate = sheetData.sort_values(by='漏失率', ascending=False) # 按列排序

# pandas.to_numeric(df1["ygyxje"], errors=0)     # 更改数据类型

# a_list = [1, 2, 3, 4, 5]
# b_list = [1, 4, 5]
# ret_list = list(set(a_list)^set(b_list))   # 两个列表求差集^   交集&   并集|
# data1 = sheetData[~sheetData['findingno'].isin(dataInSql)]  # isin判断数据是否在列表内，取反加~
# import os
# path = r'D:/资料汇总/python自学/python脚本语句'
# for filename in os.listdir(path):
# print(os.path.join(path, filename))
# import os

# print '***获取当前目录***'
# print os.getcwd()
# print os.path.abspath(os.path.dirname(__file__))

# print '***获取上级目录***'
# print os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
# print os.path.abspath(os.path.dirname(os.getcwd()))
# print os.path.abspath(os.path.join(os.getcwd(), ".."))

# print '***获取上上级目录***'
# print os.path.abspath(os.path.join(os.getcwd(), "../.."))

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 1.打开工作簿
wb = openpyxl.load_workbook('files/school.xlsx')
sheet = wb.active

# 2.设置单元格字体样式
"""
Font(
    name=None,      # 字体名，可以用字体名字的字符串
    strike=None,    # 删除线，True/False
    color=None,     # 文字颜色
    size=None,      # 字号
    bold=None,      # 加粗, True/False
    italic=None,    # 倾斜，Tue/False
    underline=None # 下划线, 'singleAccounting', 'double', 'single', 'doubleAccounting'
)
"""
# 1) 创建字体对象
font1 = Font(
    size=20,
    italic=True,
    color='ff0000',
    bold=True,
    strike=True
)
# 2) 设置指定单元格的字体
# 单元格对象.font = 字体对象
sheet['B2'].font = font1

# 3. 设置单元格填充样式
"""
PatternFill(
	fill_type=None,		# 设置填充样式: 'darkGrid', 'darkTrellis', 'darkHorizontal', 'darkGray', 'lightDown', 'lightGray', 'solid', 'lightGrid', 'gray125', 'lightHorizontal', 'lightTrellis', 'darkDown', 'mediumGray', 'gray0625', 'darkUp', 'darkVertical', 'lightVertical', 'lightUp'
	start_color=None	# 设置填充颜色
)
"""
# 1） 设置填充对象
fill = PatternFill(
    fill_type='solid',
    start_color='ffff00'
)
# 2）设置单元格的填充样式
# 单元格对象.fill = 填充对象
sheet['B2'].fill = fill

# 4. 设置单元格对齐样式
# 1）创建对象
al = Alignment(
    horizontal='right',  # 水平方向:center, left, right
    vertical='top'  # 垂直方向: center, top, bottom
)
# 2） 设置单元格的对齐方式
sheet['B2'].alignment = al

# 5. 设置边框样式
# 1）设置边对象（四个边的边可以是一样的也可以不同，如果不同就创建对个Side对象）
side = Side(border_style='thin', color='0000ff')
# 2) 设置边框对象
# 这儿的left、right、top、bottom表示的是边框的四个边，这儿四个边使用的是一个边对象
bd = Border(left=side, right=side, top=side, bottom=side)
# 3）设置单元格的边框
sheet['B2'].border = bd

# 6.设置单元格的宽度和高度
# 设置指定列的宽度
sheet.column_dimensions['A'].width = 20
# 设置指定行的高度
sheet.row_dimensions[1].height = 45

# 7. 保存
wb.save('files/school.xlsx')
