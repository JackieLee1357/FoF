#!/E:\PythonProjects\pythonProject2 python3.8
# -*- coding: utf-8 -*-
# ---
# @Software: PyCharm
# @File: testMail.py
# @Author: Jackie Lee
# @Institution: Wuxi, Jiangsu, China
# @E-mail: Yuan_li5928@jabil.com
# @Site:
# @Time: 8月 06, 2021
# ---

import base64
import configparser
import datetime
import os
import smtplib
import sys
import time
import traceback
import pandas
import sqlalchemy
import sqlalchemy.sql.default_comparator
import win32com.client
import win32com.client as win32
from PIL import ImageGrab
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from importlib import reload
from openpyxl import load_workbook


stdi, stdo, stde = sys.stdin, sys.stdout, sys.stderr
reload(sys)  # 通过import引用进来时,setdefaultencoding函数在被系统调用后被删除了，所以必须reload一次
sys.stdin, sys.stdout, sys.stderr = stdi, stdo, stde


class SendMail(object):
    def __init__(self, content=None,
                 image=None, file=None):

        '''
               :param recv: 收件人，多个要传list ['a@qq.com','b@qq.com]
               :param content: 邮件正文
               :param image: 图片路径，绝对路径，默认为无图片
               :param file: 附件路径，如果不在当前目录下，要写绝对路径，默认没有附件
        '''

        self.recv = people_email  # 收件人，多个要传list ['a @ qq.com','b @ qq.com]
        self.cc_email = cc_email  # cc人员
        self.title = 'Trace每日漏失报告'  # 邮件标题
        self.content = content  # 邮件正文
        self.image = image  # 图片路径（绝对路径）
        self.file = file  # 图片路径（绝对路径）
        self.message = MIMEMultipart()  # 构造一个MIMEMultipart对象代表邮件本身

        # 添加文件到附件
        if self.file:
            file_name = os.path.split(self.file)[-1]  # 只取文件名，不取路径
            try:
                f = open(self.file, 'rb').read()
            except Exception as e:
                traceback.print_exc()
            else:
                att = MIMEText(f, "base64", "utf-8")
                att["Content-Type"] = 'application/octet-stream'
                # base64.b64encode(file_name.encode()).decode()
                new_file_name = '=?utf-8?b?' + base64.b64encode(file_name.encode()).decode() + '?='
                # 处理文件名为中文名的文件
                att["Content-Disposition"] = 'attachment; filename="%s"' % (new_file_name)
                self.message.attach(att)

        # 添加图片到附件
        if self.image:
            try:
                with open(self.image, 'rb') as f:
                    # 将图片显示在邮件正文中
                    msgimage = MIMEImage(f.read())
                    msgimage.add_header('Content-ID', '<image1>')  # 指定文件的Content-ID,<img>,在HTML中图片src将用到
                    self.message.attach(msgimage)
            except Exception as e:
                traceback.print_exc()

    def send_mail(self):
        mailTitle = 'Trace每日漏失报告'
        from_addr = "OE-FoF_supportTeam@jabil.com"
        receivers = self.recv
        cc = self.cc_email
        to_addrs = receivers + cc
        smtpServer = r'CORIMC04'
        commonPort = 587
        smtp = smtplib.SMTP("{}:{}".format(smtpServer, commonPort))
        self.message.attach(MIMEText(self.content, 'html', 'utf-8'))  # 正文内容   plain代表纯文本,html代表支持html文本
        self.message['From'] = from_addr  # 发件人
        self.message['To'] = ','.join(receivers)  # 收件人
        self.message['Subject'] = Header(mailTitle, 'utf-8').encode()  # 邮件标题
        self.message['Cc'] = ','.join(cc)  # cc
        # 发送邮件服务器的对象
        print('收件人为------------------')
        print(receivers)
        print('抄送------------------')
        print(cc)
        try:
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addrs, msg=self.message.as_string())
            print("邮件发送成功！")
            pass
        except Exception as e:
            resultCode = 0
            traceback.print_exc()
        else:
            resultCode = 1
        smtp.quit()
        return resultCode


def readExcel(pathName):
    # 读取数据源excel，进行数据处理，并返回DataFrame
    sheetKeys = pandas.read_excel(pathName, sheet_name=None)  # 读取excel
    pandas.set_option('display.max_columns', None)  # 显示所有列
    # print(list(sheetKeys))   #显示所有sheet名称
    sheetFACA = pandas.read_excel(pathName, sheet_name=list(sheetKeys).index('FACA汇总'))
    sheetFACA.columns = range(sheetFACA.shape[1])  # shape[1]  列数
    sheetFACA.drop(sheetFACA.columns[list(range(11, sheetFACA.shape[1]))], axis=1, inplace=True)  # 去掉多余数据
    sheetData = pandas.read_excel(pathName, sheet_name=list(sheetKeys).index('数据'))
    sheetData.columns = range(sheetData.shape[1])
    index = list(sheetData.values[1]).index(dateDays)  # 当前日期的index
    print('--------')
    sheetData.drop(sheetData.columns[list(range(int(index) + 5, sheetData.shape[1]))], axis=1, inplace=True)  # 去掉多余数据
    sheetData.drop(sheetData.columns[list(range(6, int(index)))], axis=1, inplace=True)  # 去掉多余数据
    for i in range(len(sheetData)):
        if sheetData.loc[i, 0] == '制程':
            sheetData = sheetData.loc[i + 2:]
            break
    for i in range(len(sheetFACA)):
        if sheetFACA.loc[i, 1] == '日期':
            sheetFACA = sheetFACA.loc[i + 1:]
            break
    sheetFACA.reset_index(drop=True, inplace=True)  # 行号重新排序
    for i in range(len(sheetFACA)):
        sheetFACA.loc[i, 0] = i + 3  # 获取itemID
        if sheetFACA.loc[i, 1] != dateNow:      # 去掉非今日数据
            sheetFACA = sheetFACA.drop(i)
    df = pandas.merge(left=sheetFACA, right=sheetData, how='left', left_on=3, right_on=1)  # inner join
    dfColumns = ['呈现站点', 'item_id', '日期', '被卡控站点', '呈现站点1', '问题描述', 'dri', 'fa', 'ca', '改善开始日期',
                 '预计完成日期', '状态', '制程', '呈现站点2', '被卡控站点2', '漏失目标', 'dri2', '被卡控站点3', '总数', '漏失数量', '漏失率',
                 '问题点描述', '问题类型']
    df.columns = dfColumns
    sheetData.loc[0:, '日期'] = dateNow
    sheetDataTableName = ['制程', '站别', '被卡控站点', '漏失目标', 'dri', '呈现工站', '总数', '漏失数量', '漏失率',
                          '问题点描述', '问题类型', '日期']
    sheetData.columns = sheetDataTableName
    df = df.drop_duplicates()  # 去除重复数据
    print('源数据读取完毕')
    print('------------')
    return [df, sheetData]


def conPGSQL(sqlData):  # 链接PG数据库，查询已处理的时段
    db_url = 'postgresql+psycopg2://OE_User:JGP123456@CNWXIM0WINSVC01:5438/Trace_T1'
    engine = sqlalchemy.create_engine(db_url)
    connection = engine.raw_connection()
    cursor = connection.cursor()
    cursor.execute(sqlData)
    row = cursor.fetchall()
    # row =  [''.join(i) for i in row]    #元组转化为列表
    connection.commit()
    cursor.close()
    engine.dispose()
    print('------------')
    return row


def getMailData(sqlData, tableColumns):
    # pandas.DataFrame(dataFrame).to_excel('C:\ActionTracker\data1.xlsx', sheet_name='summery', index=False, header=True)
    mailInfoTable = pandas.DataFrame(conPGSQL(sqlData))  # 从数据库获取需回复FACA信息
    mailInfoTable.columns = tableColumns
    for i in range(len(mailInfoTable.loc[0:, '漏失率'])):
        mailInfoTable.loc[i, '漏失率'] = '{:.2%}'.format(float(mailInfoTable.loc[i, '漏失率']))  # 格式化输出%数据
    print('------------')
    return mailInfoTable


def dataSplit(dataList):
    # 把数据分成三部分，FACA表和TraceMissing表
    dataFrame = dataList[0]
    # pandas.DataFrame(dataFrame).to_excel('E:\OneDrive\OneDrive - Jabil\ActionTracker\data1.xlsx', sheet_name='summery',
    #                                      index=False, header=True)
    traceMissingTableName = ['item_id', '日期', '被卡控站点', '呈现站点', '漏失数量', '漏失率', 'dri']
    if dataFrame.empty:  # 判断DataFrame是否为空
        print('今日数据未更新，请更新数据！')
        exit(0)
    traceMissingTable = dataFrame.loc[0:, traceMissingTableName]  # 选取特定列
    FACATableName = ['item_id', '日期', 'fa', 'ca']
    FACATable = dataFrame.loc[0:, FACATableName]  # 选取特定列
    FACATable.loc[0:, 'id'] = FACATable.index  # 新增一列
    FACATable.loc[0:, 'project'] = 'TraceMissing'  # 新增一列
    FACATable.loc[0:, 'returndri'] = '待回复'  # 新增一列
    FACATable.loc[0:, 'returntime'] = '待回复'  # 新增一列
    FACATable = FACATable.loc[0:,
                ['id', 'item_id', 'project', '日期', 'fa', 'ca', 'returndri', 'returntime']]  # 选取特定列,重新排序
    FACATable.columns = ['id', 'item_id', 'project', 'eventtime', 'fa', 'ca', 'returndri', 'returntime']
    FACATable = FACATable.fillna(value='待回复')  # 替换NaN值
    FACATable.loc[0:, 'returntime'] = None
    # FACATable.style
    # traceMissingTable.style
    print('数据处理完成')
    print('------------')
    return [FACATable, traceMissingTable, dataList[1]]


def writeToSheets(pathName, sheets, dfs):
    try:
        workbook1 = load_workbook(pathName)
    except:
        print('excel不存在，重新创建！')
        writeToSheets2(pathName, sheets, dfs)
        workbook1 = load_workbook(pathName)
    sheet1 = workbook1[sheets[0]]
    df1 = dfs[0]
    excelHeader = list(df1.columns)
    for row in range(len(df1)):
        for index in range(len(df1.loc[row])):
            sheet1.cell(row + 2, index + 2).value = df1.loc[row, list(df1.columns)[index]]  # 更改已经存在的数据
            # if row == 0:
            #     sheet1.cell(row + 1, index + 2).value = excelHeader[index]  # 更改已经存在的数据
    for i in range(row + 3, 50):
        for j in range(2, len(df1.loc[row]) + 2):
            sheet1.cell(i, j).value = ''
    # sheet2 = workbook1[sheets[1]]
    # df2 = dfs[1]
    # for row in range(len(df2)):
    #     for index in range(len(df2.loc[row])):
    #         sheet2.cell(row + 2, index + 2).value = df2.loc[row, list(df2.columns)[index]]
    # for i in range(row + 3, 50):
    #     for j in range(2, len(df2.loc[row]) + 2):
    #         sheet1.cell(i, j).value = ''
    workbook1.save(pathName)  # 保存修改
    print('excel数据修改完毕')
    print('------------')


def writeToSheets2(pathName, sheets, dfs):
    # 数据写入excel中对应的sheet
    # 打开excel
    writer = pandas.ExcelWriter(pathName)
    # sheets是要写入的excel工作簿名称列表
    for i in range(len(sheets)):
        dfs[i].to_excel(writer, sheet_name=sheets[i])
        # 保存writer中的数据至excel
    writer.save()
    print('数据已写入EXCEL')
    print('------------')


def insertIntoSql(frames, tableNames):
    # 数据插入PG sql
    db_url = 'postgresql+psycopg2://OE_User:JGP123456@CNWXIM0WINSVC01:5438/Trace_T1'
    engine = sqlalchemy.create_engine(db_url)
    connection = engine.raw_connection()
    for i in range(len(frames)):
        frames[i].to_sql(tableNames[i], engine, index=False, if_exists='append')
    engine.dispose()
    print('数据写入PG SQL成功')
    print('------------')


def sentOutlookMail(dailyReport, people_email):
    mail_content = 'Trace每日漏失报告'
    smtpServer = r'CORIMC04'
    commonPort = 587
    smtp = smtplib.SMTP("{}:{}".format(smtpServer, commonPort))
    message = MIMEMultipart()
    msg = MIMEText(dailyReport, "html", 'utf-8')
    msg["Subject"] = Header(mail_content, 'utf-8').encode()
    from_addr = "OE-FoF_supportTeam@jabil.com"
    receivers = people_email
    receivers = list(receivers)
    msg["To"] = ",".join(receivers)
    msg["from"] = from_addr
    smtp.sendmail(from_addr=from_addr, to_addrs=receivers, msg=msg.as_string())
    smtp.quit()
    print('邮件发送成功')
    print('------------')


def changeStyle(mailInfo, mailTime):
    with open(imagePath, "rb") as f:   # 图片转为base64格式
        base64_data = base64.b64encode(f.read())
        imageBase64 = base64_data.decode()

    head = \
        """
        <head>
            <meta charset="utf-8">
            <STYLE TYPE="text/css" MEDIA=screen>
                table
            {border-collapse:collapse;}
            table, td,tr,th
            {border:1px solid black;padding: 15px}
            </STYLE>
        </head>
        """
    body = \
        """
       <body>
        <div align="left",font-family='SimSun',"font-size"="15px">
           <p>Dear Sirs:</p>
           <br>
           <p>以下为Trace每日漏失报告,请查阅！</p>
           <p>发送日期：{sh_sys}</p>
           <p>{dataFrame}</p>
           <p><img src="data1:image/jpeg;base64,{imageBase64}" alt=" " width="1080"></p>   
           <br>
           <a href="https://apps.powerapps.com/play/220d9b3e-6604-4257-9eb3-79e06c05f5b2?tenantId=bc876b21-f134-4c12-
           a265-8ed26b7f0f3b&source=portal&screenColor=rgba(0%2C%20176%2C%20240%2C%201)
           "target="_blank">请用手机登录Power APPs，搜索Trace漏失应用，回复FA/CA！或点击此链接登录网页版Power APPs回复FA/CA！</a>
           <br>
           <br>
           <br>
           <a href="http://cnwgpm0pbi01/Reports/powerbi/WX_MetalSME/official/TEST/Trace%E7%94%9F%E4%BA%A7%E6%BC%8F%E5%A4%B1"
           target="_blank">更多Trace漏失数据请查阅PowerBi，请用厂内Jabil网络登录！</a>
           <br>
           <br>
           <p>Best Regards!</p>
           <p>Send by: OE-FoF</p>
           <p>自动邮件,请勿答复</p>
        </div>
    </body>
    """.format(dataFrame=mailInfo, sh_sys=str(mailTime)[0:10], imageBase64=imageBase64)
    mailText = "<html>" + head + body + "</html>"
    # dataFrame = dataFrame.replace('&&&', '<br>')
    print('mail格式处理完毕')
    print('------------')
    return mailText


def changeColor(val):
    """
    :param val:元素值，表达式datafarmede中的所有值，单个传入 对单个值的字体颜色修改
    :return: val
    """
    try:
        if float(val[:-1]) > 0.05:  # 报警规则
            return 'color:red'
        else:
            return 'color:black'
    except Exception as e:
        print(e)
        return 'color:black'


def getPictureFromExcel(sourcePath1, imagePath1):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    workbook = excel.Workbooks.Open(sourcePath1)
    workSheet = workbook.Worksheets("生产异常汇总")

    for i, shape in enumerate(workSheet.Shapes):
        if shape.Name.startswith('Picture'):
            shape.Copy()
            image = ImageGrab.grabclipboard()
            image.convert('RGB').save(imagePath1)
    workbook.Save()
    workbook.Close()
    excel.Quit()
    print("图片读取成功~")


def openExcel(filename):
    # 模拟电脑打开excel文件
    xlApp = win32com.client.Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()
    xlApp.Quit()
    print("excel打开保存完毕~")


if __name__ == '__main__':
    path0 = os.path.abspath(os.path.dirname(sys.argv[0]))  # 获取当前执行文件夹路径
    path0 = path0+"/actionTracker.ini"
    config = configparser.ConfigParser()
    try:
        config.read(path0, encoding="utf-8-sig")  # 导出配置文件,解决文档格式为utf-8-bom问题
    except Exception as e:
        print(e)
        config.read(path0, encoding="utf-8")  # 导出配置文件
    people_email = config.get("messages", "people_email")  # 接收邮件人员
    cc_email = config.get("messages", "cc_email")  # cc邮件人员
    people_email = people_email.split(",")   # 字符串根据","进行分割，转换为list
    cc_email = cc_email.split(",")
    sourcePath = config.get("messages", "sourcePath")  # 数据源地址
    outputPath = config.get("messages", "outputPath")  # 处理后文件路径
    dayAdd = int(config.get("messages", "dayAdd"))  # 取之前多少天的日期
    reMail = config.get("messages", "remail")  # 判断是否允许重复发邮件， 1为允许
    imagePath = config.get("messages", "imagePath")  # 图片路径
    SQLTableName = ['actiontracker_faca', 'actiontracker_tracemissing', 'actiontracker_data']  # 数据库表名称

    date0 = datetime.datetime.strptime("1900-01-01", "%Y-%m-%d")
    dateNow = datetime.datetime.strptime(
        (datetime.datetime.now() + datetime.timedelta(days=dayAdd)).strftime("%Y-%m-%d"),
        "%Y-%m-%d")
    dateDays = (dateNow - date0).days + 2
    print(f'获取的数据日期为：{dateNow}, 数字格式为：{dateDays}')
    sql = 'SELECT DISTINCT 日期 FROM actiontracker_tracemissing'
    dates = conPGSQL(sql)
    dates = [''.join(i) for i in dates]  # 元组转化为列表
    isRun = (f'{dateNow}' in dates)  # 判断今日数据是否已运行，True为已运行数据， False为未运行

    dfList = readExcel(sourcePath)  # 1.从excel读取数据源
    getPictureFromExcel(sourcePath, imagePath)       # 从excel读取图片
    dataList = dataSplit(dfList)  # 2.对数据进行分析，拆成两个表
    print('FACA数据库数据为-----')
    print(dataList[0])
    if not isRun:  # 防止重复存入数据库
        insertIntoSql(dataList, SQLTableName)  # 3.数据存入PGSQL

    facaInfoTableName = ['item_id', 'project', '日期', '被卡控站点', '呈现站点', '漏失数量', '漏失率', 'FA', 'CA', 'DRI', 'returndri',
                         'returntime']
    getFacaSqlData = f""" SELECT DISTINCT
                    A.item_id, project, A.日期, 被卡控站点, 呈现站点, 漏失数量, 漏失率, FA, CA, DRI, returndri, returntime
                    FROM actiontracker_tracemissing A INNER JOIN actiontracker_faca B
                    ON A.item_id = B.item_id
                    WHERE A.日期 = '{dateNow}' and b.eventtime = '{dateNow}'
                    ORDER BY A.日期 DESC """
    facaInfo = getMailData(getFacaSqlData, facaInfoTableName)  # 4. 从数据库获取excel Date信息pd
    dataList2 = [facaInfo, dataList[1]]
    print('FACA表数据为-----')
    print(facaInfo)
    if not isRun:  # 防止重复写入excel   NOT
        writeToSheets(outputPath, ['FACA', 'TraceMissing'], dataList2)  # 5. 数据写入excel-Date
        # exit(0)
    mailTime = datetime.datetime.now().strftime("%Y-%m-%d")
    mailTime = mailTime[0:10]
    mailInfoTableName = ['日期', '被卡控站点', '呈现站点', '漏失数量', '漏失率', 'FA', 'CA', 'DRI']
    getMailSqlData = f""" SELECT DISTINCT
                A.日期, 被卡控站点, 呈现站点, 漏失数量, 漏失率, FA, CA, DRI
                FROM actiontracker_tracemissing A INNER JOIN actiontracker_faca B
                ON A.item_id = B.item_id
                WHERE A.日期 = '{dateNow}' and b.eventtime = '{dateNow}' 
                ORDER BY A.日期 DESC """
    mailInfo = getMailData(getMailSqlData, mailInfoTableName)  # 6. 从数据库获取邮件信息pd
    styles = [dict(selector='td', props=[("text-align", "center")]),
              dict(selector="th", props=[("font-size", "15px"), ("text-align", "center")]),
              dict(selector="caption", props=[("caption-side", "top")])]
    mailInfo = mailInfo.style.set_table_styles(styles).applymap(changeColor, subset=['漏失率']).hide_index().render()
    mailInfo = changeStyle(mailInfo, mailTime)
    openExcel(outputPath)     # 模拟电脑打开excel
    if isRun & (reMail != '1'):  # 防止重复运行重复存入数据库重复发邮件
        print('当天数据已运行')
        time.sleep(5)  # 暂停5秒
        exit(0)
    m = SendMail(content=mailInfo, image=None)
    m.send_mail()  # 7. 发Trace漏失邮件
    time.sleep(5)  # 暂停5秒
