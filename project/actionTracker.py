#!C:\ProgramData\Anaconda3 python3.
# -*- coding: utf-8 -*-
# ---
# @Software: PyCharm
# @File: ppDowntime.py
# @Author: Jackie Lee
# @Institution: Wuxi, Jiangsu, China
# @E-mail: Yuan_li5928@jabil.com
# @Site:
# @Time: 8月 02, 2021
# ---

from importlib import reload
import jinja2
from io import StringIO
from email.header import Header
from email.mime.text import MIMEText
import smtplib
import datetime
import configparser
import sqlalchemy.sql.default_comparator
import time
import pandas
import sqlalchemy
from openpyxl import load_workbook
import sys  # 这里只是一个对sys的引用，只能reload才能进行重新加载

stdi, stdo, stde = sys.stdin, sys.stdout, sys.stderr
reload(sys)  # 通过import引用进来时,setdefaultencoding函数在被系统调用后被删除了，所以必须reload一次
sys.stdin, sys.stdout, sys.stderr = stdi, stdo, stde


def readExcel(pathName):
    # 读取数据源excel，进行数据处理，并返回DataFrame
    sheetKeys = pandas.read_excel(pathName, sheet_name=None)  # 读取excel
    pandas.set_option('display.max_columns', None)  # 显示所有列
    # print(list(sheetKeys))   #显示所有sheet名称
    sheetFACA = pandas.read_excel(pathName, sheet_name=list(sheetKeys).index('FACA汇总'))
    sheetFACA.columns = range(sheetFACA.shape[1])  # shape[1]  列数
    sheetFACA.drop(sheetFACA.columns[list(range(11, sheetFACA.shape[1]))], axis=1, inplace=True)  # 去掉多余数据
    sheetData = pandas.read_excel(pathName, sheet_name=list(sheetKeys).index('数据'))
    sheetData.columns = range(sheetData.shape[1])
    index = list(sheetData.values[1]).index(date)  # 当前日期的index
    print('--------')
    sheetData.drop(sheetData.columns[list(range(int(index) + 5, sheetData.shape[1]))], axis=1, inplace=True)  # 去掉多余数据
    sheetData.drop(sheetData.columns[list(range(6, int(index)))], axis=1, inplace=True)  # 去掉多余数据
    for i in range(len(sheetData)):
        if sheetData.loc[i, 0] == '制程':
            sheetData = sheetData.loc[i + 2:]
            break
    for i in range(len(sheetFACA)):
        if sheetFACA.loc[i, 1] == '日期':
            sheetFACA = sheetFACA.loc[i + 1:]
            break
    sheetFACA.reset_index(drop=True, inplace=True)  # 行号重新排序
    for i in range(len(sheetFACA)):
        sheetFACA.loc[i, 0] = i + 3  # 获取itemID
        if sheetFACA.loc[i, 1] < dateNow:
            sheetFACA = sheetFACA.drop(i)
    df = pandas.merge(left=sheetFACA, right=sheetData, how='left', left_on=3, right_on=1)  # inner join
    dfColumns = ['呈现站点', 'item_id', '日期', '呈现站点1', '被卡控站点', '问题描述', 'dri', 'fa', 'ca', '改善开始日期',
                 '预计完成日期', '状态', '制程', '站别', '呈现站点3', '漏失目标', 'dri2', '呈现站点2', '总数', '漏失数量', '漏失率',
                 '问题点描述', '问题类型']
    df.columns = dfColumns
    sheetData.loc[0:, '日期'] = dateNow
    sheetDataTableName = ['制程', '站别', '被卡控站点', '漏失目标', 'dri', '呈现工站', '总数', '漏失数量', '漏失率',
                          '问题点描述', '问题类型', '日期']
    sheetData.columns = sheetDataTableName
    df = df.drop_duplicates()  # 去除重复数据
    print('源数据读取完毕')
    print('------------')
    return [df, sheetData]


def conPGSQL(sqlData):  # 链接PG数据库，查询已处理的时段
    db_url = 'postgresql+psycopg2://OE_User:JGP123456@CNWXIM0WINSVC01:5438/Trace_T1'
    engine = sqlalchemy.create_engine(db_url)
    connection = engine.raw_connection()
    cursor = connection.cursor()
    cursor.execute(sqlData)
    row = cursor.fetchall()
    # row =  [''.join(i) for i in row]    #元组转化为列表
    connection.commit()
    cursor.close()
    engine.dispose()
    print('------------')
    return row


def getMailData(sqlData, tableColumns):
    # pandas.DataFrame(dataFrame).to_excel('C:\ActionTracker\data1.xlsx', sheet_name='summery', index=False, header=True)
    mailInfoTable = pandas.DataFrame(conPGSQL(sqlData))  # 从数据库获取需回复FACA信息
    mailInfoTable.columns = tableColumns
    for i in range(len(mailInfoTable.loc[0:, '漏失率'])):
        mailInfoTable.loc[i, '漏失率'] = '{:.1%}'.format(float(mailInfoTable.loc[i, '漏失率']) / 100)  # 格式化输出%数据
    print('------------')
    return mailInfoTable


def dataSplit(dataList):
    # 把数据分成三部分，FACA表和TraceMissing表
    dataFrame = dataList[0]
    pandas.DataFrame(dataFrame).to_excel('E:\OneDrive\OneDrive - Jabil\ActionTracker\data1.xlsx', sheet_name='summery',
                                         index=False, header=True)
    traceMissingTableName = ['item_id', '日期', '被卡控站点', '呈现站点', '漏失数量', '漏失率', 'dri']
    if dataFrame.empty:  # 判断DataFrame是否为空
        print('今日数据未更新，请更新数据！')
        exit(0)
    traceMissingTable = dataFrame.loc[0:, traceMissingTableName]  # 选取特定列
    FACATableName = ['item_id', '日期', 'fa', 'ca']
    FACATable = dataFrame.loc[0:, FACATableName]  # 选取特定列
    FACATable.loc[0:, 'id'] = FACATable.index  # 新增一列
    FACATable.loc[0:, 'project'] = 'TraceMissing'  # 新增一列
    FACATable.loc[0:, 'returndri'] = '待回复'  # 新增一列
    FACATable.loc[0:, 'returntime'] = '待回复'  # 新增一列
    FACATable = FACATable.loc[0:,
                ['id', 'item_id', 'project', '日期', 'fa', 'ca', 'returndri', 'returntime']]  # 选取特定列,重新排序
    FACATable.columns = ['id', 'item_id', 'project', 'eventtime', 'fa', 'ca', 'returndri', 'returntime']
    FACATable = FACATable.fillna(value='待回复')  # 替换NaN值
    FACATable.loc[0:, 'returntime'] = None
    FACATable.style
    traceMissingTable.style
    print('数据处理完成')
    print('------------')
    return [FACATable, traceMissingTable, dataList[1]]


def writeToSheets(pathName, sheets, dfs):
    try:
        workbook1 = load_workbook(pathName)
    except:
        print('excel不存在，重新创建！')
        writeToSheets2(pathName, sheets, dfs)
        workbook1 = load_workbook(pathName)
    sheet1 = workbook1[sheets[0]]
    df1 = dfs[0]
    for row in range(len(df1)):
        for index in range(len(df1.loc[row])):
            sheet1.cell(row + 2, index + 2).value = df1.loc[row, list(df1.columns)[index]]  # 更改已经存在的测试数据
    for i in range(row + 3, 50):
        for j in range(2, len(df1.loc[row]) + 2):
            sheet1.cell(i, j).value = ''
    sheet2 = workbook1[sheets[1]]
    df2 = dfs[1]
    for row in range(len(df2)):
        for index in range(len(df2.loc[row])):
            sheet2.cell(row + 2, index + 2).value = df2.loc[row, list(df2.columns)[index]]
    for i in range(row + 3, 50):
        for j in range(2, len(df2.loc[row]) + 2):
            sheet1.cell(i, j).value = ''
    workbook1.save(pathName)  # 保存修改
    print('excel数据修改完毕')
    print('------------')


def writeToSheets2(pathName, sheets, dfs):
    # 数据写入excel中对应的sheet
    # 打开excel
    writer = pandas.ExcelWriter(pathName)
    # sheets是要写入的excel工作簿名称列表
    for i in range(len(sheets)):
        dfs[i].to_excel(writer, sheet_name=sheets[i])
        # 保存writer中的数据至excel
    writer.save()
    print('数据已写入EXCEL')
    print('------------')


def insertIntoSql(frames, tableNames):
    # 数据插入PG sql
    db_url = 'postgresql+psycopg2://OE_User:JGP123456@CNWXIM0WINSVC01:5438/Trace_T1'
    engine = sqlalchemy.create_engine(db_url)
    connection = engine.raw_connection()
    for i in range(len(frames)):
        frames[i].to_sql(tableNames[i], engine, index=False, if_exists='append')
    engine.dispose()
    print('数据写入PG SQL成功')
    print('------------')


def sentOutlookMail(dailyReport, people_email):
    mail_content = 'Trace每日漏失报告'
    smtpServer = r'CORIMC04'
    commonPort = 587
    smtp = smtplib.SMTP("{}:{}".format(smtpServer, commonPort))
    msg = MIMEText(dailyReport, "html", 'utf-8')
    msg["Subject"] = Header(mail_content, 'utf-8').encode()
    from_addr = "OE-FoF_supportTeam@jabil.com"
    receivers = people_email
    receivers = list(receivers)
    msg["To"] = ",".join(receivers)
    msg["from"] = from_addr
    smtp.sendmail(from_addr=from_addr, to_addrs=receivers, msg=msg.as_string())
    smtp.quit()
    print('邮件发送成功')
    print('------------')


def changeStyle(mailInfo, mailTime):
    head = \
        """
        <head>
            <meta charset="utf-8">
            <STYLE TYPE="text/css" MEDIA=screen>
                table
            {border-collapse:collapse;}
            table, td,tr,th
            {border:1px solid black;padding: 15px}
            </STYLE>
        </head>
        """
    body = \
        """
       <body>
        <div align="left",font-family='SimSun',"font-size"="15px">
           <p>Dear Sirs:</p>
           <p>.    </p>
           <p>以下为Trace每日漏失报告,请查阅！</p>
           <p>发送日期：{sh_sys}</p>
           <p>{dataFrame}</p>
           <p>.    </p>
           <br/>
           <a href="https://apps.powerapps.com/play/220d9b3e-6604-4257-9eb3-79e06c05f5b2?tenantId=bc876b21-f134-4c12-a265-8ed26b7f0f3b&source=portal&screenColor=rgba%280%2C+176%2C+240%2C+1%29&skipAppMetadata=true"target="_blank">请用手机或者电脑登录Power APPs,回复FA/CA！</a>
           <p>.    </p>
           <br/>
           <a href="http://cnwgpm0pbi01/Reports/powerbi/WX_MetalSME/official/TEST/%E7%94%9F%E4%BA%A7%E6%BC%8F%E5%A4%B1"target="_blank">更多Trace漏失数据请查阅PowerBi</a>
           <p>.    </p>
           <p>Best Regards!</p>
           <p>Send by: OE-FoF</p>
           <p>自动邮件,请勿答复</p>    
        </div>
    </body>
    """.format(dataFrame=mailInfo, sh_sys=str(mailTime)[0:10])
    mailText = "<html>" + head + body + "</html>"
    #dataFrame = dataFrame.replace('&&&', '<br>')
    print('mail格式处理完毕')
    print('------------')
    return mailText


def changeColor(val):
    """
    :param val:元素值，表达式datafarmede中的所有值，单个传入 对单个值的字体颜色修改
    :return: val
    """
    try:
        if float(val[:-1]) > 0.05:  # 报警规则
            return 'color:red'
        else:
            return 'color:black'
    except Exception as e:
        print(e)
        return 'color:black'


if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read('C:\ActionTracker\FACA.ini', encoding="utf-8")  # 导出配置文件
    people_email = config.get("messages", "people_email")  # 接收邮件人员
    sourcePath = config.get("messages", "sourcePath")      # 数据源地址
    outputPath = config.get("messages", "outputPath")      # 处理后文件网址
    dayAdd = int(config.get("messages", "dayAdd"))         # 取之前多少天的日期
    reMail = config.get("messages", "remail")              # 判断是否允许重复发邮件， 1为允许
    SQLTableName = ['actiontracker_faca', 'actiontracker_tracemissing', 'actiontracker_data']   #数据库表名称

    date0 = datetime.datetime.strptime("1900-01-01", "%Y-%m-%d")
    dateNow = datetime.datetime.strptime(
        (datetime.datetime.now() + datetime.timedelta(days=dayAdd)).strftime("%Y-%m-%d"),
        "%Y-%m-%d")
    date = (datetime.datetime.now() + datetime.timedelta(days=dayAdd) - date0).days
    sql = 'SELECT DISTINCT 日期 FROM actiontracker_tracemissing'
    dates = conPGSQL(sql)
    dates = [''.join(i) for i in dates]  # 元组转化为列表
    isRun = (f'{dateNow}' in dates)  # 判断今日数据是否已运行，True为已运行数据， False为未运行

    dfList = readExcel(sourcePath)  # 1.从excel读取数据源
    dataList = dataSplit(dfList)  # 2.对数据进行分析，拆成两个表
    if not isRun:  # 防止重复存入数据库
        insertIntoSql(dataList, SQLTableName)  # 3.数据存入PGSQL

    facaInfoTableName = ['item_id', 'project', '日期', '被卡控站点', '呈现站点', '漏失数量', '漏失率', 'FA', 'CA', 'DRI', 'returndri',
                         'returntime']
    getFacaSqlData = f""" SELECT DISTINCT
                    A.item_id, project, A.日期, 被卡控站点, 呈现站点, 漏失数量, 漏失率, FA, CA, DRI, returndri, returntime
                    FROM actiontracker_tracemissing A INNER JOIN actiontracker_faca B
                    ON A.item_id = B.item_id
                    WHERE A.日期 > '{dateNow.strftime("%Y-%m-%d")[:10]}' 
                    ORDER BY A.日期 DESC """
    facaInfo = getMailData(getFacaSqlData, facaInfoTableName)  # 4. 从数据库获取excel Date信息pd
    dataList2 = [facaInfo, dataList[1]]
    if not isRun:  # 防止重复写入excel
        writeToSheets(outputPath, ['FACA', 'TraceMissing'], dataList2)  # 5. 数据写入excel-Date
    people_email = people_email.split(",")
    mailTime = datetime.datetime.now().strftime("%Y-%m-%d")
    mailTime = mailTime[0:10]
    mailInfoTableName = ['日期', '被卡控站点', '呈现站点', '漏失数量', '漏失率', 'FA', 'CA', 'DRI']
    getMailSqlData = f""" SELECT DISTINCT
                A.日期, 被卡控站点, 呈现站点, 漏失数量, 漏失率, FA, CA, DRI
                FROM actiontracker_tracemissing A INNER JOIN actiontracker_faca B
                ON A.item_id = B.item_id
                WHERE A.日期 > '{dateNow.strftime("%Y-%m-%d")[:10]}' 
                ORDER BY A.日期 DESC """
    time.sleep(3)
    mailInfo = getMailData(getMailSqlData, mailInfoTableName)  # 6. 从数据库获取邮件信息pd
    styles = [dict(selector='td', props=[("text-align", "center")]),
              dict(selector="th", props=[("font-size", "15px"), ("text-align", "center")]),
              dict(selector="caption", props=[("caption-side", "top")])]
    mailInfo = mailInfo.style.set_table_styles(styles).applymap(changeColor, subset=['漏失率']).hide_index().render()
    mailInfo = changeStyle(mailInfo, mailTime)
    if isRun & (reMail != '1'):  # 防止重复运行重复存入数据库重复发邮件
        print('当天数据已运行')
        time.sleep(5)  # 暂停5秒
        exit(0)
    sentOutlookMail(mailInfo, people_email)  # 7. 发Trace漏失邮件
    time.sleep(5)                 # 暂停5秒
